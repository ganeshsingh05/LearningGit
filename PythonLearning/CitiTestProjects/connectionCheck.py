import pyodbc
import timeit
from openpyxl import Workbook

# For Calculating running time of program 
start = timeit.default_timer()
print('Process Stared:', start)

# Creating connection with MS Acess Database
AccessDriver = '{Microsoft Access Driver (*.mdb, *.accdb)}'
DatabaseFilePath = 'C:\\Users\\ganes\\Desktop\\citiTestProjects\\Books2010.accdb'

# conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\ganes\Desktop\citiTestProjects\Books2010.accdb;')

conn = pyodbc.connect(Driver = AccessDriver, DBQ = DatabaseFilePath)
print('Connection Created Successfully: ')



cursor = conn.cursor()

Query = "select * from Authors"
cursor.execute(Query)
 
   
# for row in cursor.fetchall():
#      print (row)

SampleData = cursor.fetchall()
conn.close()

#print('Data in cursor: {}'.format(SampleData))

# for item in SampleData:
#     print('Phone: {}\nFirst Name: {}\nLast Name: {}\n******'.format(item[0],item[1],item[2]))

# TODO : Create workbook and load data into that workbook 

# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = Workbook()
  
# Get workbook active sheet   
# from the active attribute.  
sheet = wb.active 


try:
    for index1, item in enumerate(SampleData):
        for index2, value in enumerate(item):
            sheet.cell(row=index1+1, column=index2+1).value=value

except ValueError:
    print('\n*++++++++++++++++++++*\nSomthing went wrong\nList data cannot be stored in single cell \nBye Bye ...! Please check your value data')
    exit()

wb.save('C:\\Users\\ganes\\Desktop\\citiTestProjects\\FinalResult.xlsx')

print('Result File Created Successfully...!')

stop = timeit.default_timer()
print('Total time taken for executing this process: {}'.format(stop - start))